from bot.bot import bot
from openpyxl import load_workbook
from discord.ext.commands import has_permissions
import time
import discord.utils
from utils.global_functions import check_s1, check_s2, check_s3


@bot.command()
@has_permissions(manage_roles=True)
async def giverole(ctx, role_name, vc_channel):
    """Command to give role to players"""
    await ctx.send('Executing !giverole')
    start_time = time.time()
    fetch = discord.utils.get(ctx.guild.voice_channels, name=vc_channel)
    user = fetch.members
    role = discord.utils.get(ctx.guild.roles, name=role_name)
    for u in user:
        await u.add_roles(role)
    await ctx.send(f'!giverole Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def test(ctx):
    """Test function"""
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    print(ws['A1'].value)


@bot.command()
@has_permissions(manage_roles=True)
async def updatescrimroles(ctx):
    """Command to update players scrim role"""
    await ctx.send('Executing !updatescrimroles')
    start_time = time.time()
    channels = ctx.guild.voice_channels
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')

    for channel in channels:
        for member in channel.members:
            if s3 not in member.roles and s2 not in member.roles and s1 not in member.roles:
                await member.add_roles(s1)
            elif s1 in member.roles:
                await member.add_roles(s2)
                await member.remove_roles(s1)
            elif s2 in member.roles:
                await member.add_roles(s3)
                await member.remove_roles(s2)

    await ctx.send(f'!updatescrimroles Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def old_updatescrimroles_2(ctx):
    """Command to update players scrim role"""
    await ctx.send('Executing !updatescrimroles')
    start_time = time.time()
    channels = ctx.guild.voice_channels
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')

    for c in channels:
        for a in c.members:
            if not check_s1(a):
                await a.add_roles(s1)
            elif not check_s2(a):
                await a.add_roles(s2)
                await a.remove_roles(s1)
            elif not check_s3(a):
                await a.add_roles(s3)
                await a.remove_roles(s2)
            else:
                pass

    await ctx.send(f'!updatescrimroles Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def old_updatescrimroles_1(ctx, vc_channel):
    """Old update scrim role function, this is not in use"""
    await ctx.send('Executing !updatescrimroles')
    start_time = time.time()
    fetch = discord.utils.get(ctx.guild.voice_channels, name=vc_channel)
    user = fetch.members
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')
    for u in user:
        for r in u.roles:
            if r == s2:
                await u.add_roles(s3)
            elif r == s1:
                await u.add_roles(s2)
            else:
                await u.add_roles(s1)

    await ctx.send(f'!updatescrimroles Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def removeallscrimroles(ctx):
    """Function to remove all scrim roles from players that had the scrim role"""
    await ctx.send('Executing !removeallscrimroles')
    start_time = time.time()
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')

    for member in s3.members:
        await member.remove_roles(s3)
    for member in s2.members:
        await member.remove_roles(s2)
    for member in s1.members:
        await member.remove_roles(s1)

    await ctx.send(f'!removeallscrimroles Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def old_removeallscrimroles_2(ctx):
    """Function to remove all scrim roles from players that had the scrim role"""
    await ctx.send('Executing !removeallscrimroles')
    start_time = time.time()
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')
    count = 1
    for row in ws.iter_rows(values_only=True):
        try:
            user = discord.utils.get(ctx.guild.members, name=str(row[2].split('#')[0]))
            await user.remove_roles(s3)
        except AttributeError:
            pass
    for row in ws.iter_rows(values_only=True):
        try:
            user = discord.utils.get(ctx.guild.members, name=str(row[1].split('#')[0]))
            await user.remove_roles(s2)
        except AttributeError:
            pass
    for row in ws.iter_rows(values_only=True):
        try:
            user = discord.utils.get(ctx.guild.members, name=str(row[0].split('#')[0]))
            await user.remove_roles(s1)
        except AttributeError:
            pass

    for row in ws.iter_rows(values_only=True):
        ws[f'A{count}'] = None
        ws[f'B{count}'] = None
        ws[f'C{count}'] = None
        count += 1

    ws[f'A1'] = 'null'
    ws[f'B1'] = 'null'
    ws[f'C1'] = 'null'
    lb.save('scrimrole.xlsx')
    await ctx.send(f'!removeallscrimroles Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def old_removeallscrimroles_1(ctx):
    """Old remove scrim role function, not in use"""
    await ctx.send('Executing !removeallscrimroles')
    start_time = time.time()
    fetch = ctx.guild.members
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')
    for f in fetch:
        for r in f.roles:
            if r == s1 or r == s2 or r == s3:
                await f.remove_roles(s1)
                await f.remove_roles(s2)
                await f.remove_roles(s3)
    await ctx.send(f'!removeallscrimroles Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def help(member):
    """Help command which shows all the available utils"""
    await member.send("```                            Help Menu\n"
                      "\nHow to use: !func(p1, p2) will be written out like !func p1 p2, for spaces use"
                      "quotes!\n\n"
                      "-!makescreenshotchannel(vc_channel, category) - Creates a text channel for all the users "
                      "inside a specified voice channel\n\n"
                      "-!deletechannels(category) - Deletes all text channels in a selected category.\n\n"
                      "-!giverole(role_name, vc_channel) - Give everyone in a specified voice chat a role of "
                      "choice.\n\n "
                      "-!updatescrimroles - Gives a user the 'scrim role'(x1, x2, x3) Example: If they "
                      "have x1 and you run the function, then x1 will be deleted and x2 will be added.\n\n"
                      "-!removeallscrimroles - Removes all scrim roles from everyone who owns the role.\n\n"
                      "** Marla bot was created by marley-EE **```")
