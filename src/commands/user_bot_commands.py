from bot.bot import bot
from openpyxl import load_workbook
import random


@bot.command()
async def duel(ctx, choice):
    """Dueling game for community interactions"""

    elduel = ['fire', 'toxic', 'wind', 'ice', 'lightning']
    lb = load_workbook(filename='duelstats.xlsx')
    ws = lb.active
    sheet_ranges = lb['Sheet']
    count = 1
    marlawins = 0

    if choice.lower() == 'fire' or choice.lower() == 'toxic' or choice.lower() == 'wind' or choice.lower() == 'ice' or \
            choice.lower() == 'lightning':
        marlapick = random.choice(elduel)
        await ctx.send(f"I choose {marlapick}!")

        if marlapick == 'fire':
            if choice.lower() == 'ice' or choice.lower() == 'toxic':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'fire':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'toxic':
            if choice.lower() == 'wind' or choice.lower() == 'lightning':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'toxic':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'wind':
            if choice.lower() == 'lightning' or choice.lower() == 'fire':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'wind':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'ice':
            if choice.lower() == 'toxic' or choice.lower() == 'wind':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'ice':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'lightning':
            if choice.lower() == 'fire' or choice.lower() == 'ice':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'lightning':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

    elif choice.lower() == 'stone':
        await ctx.send('Stone beats everything and it is too easy to use, choose a different gauntlet!')
    else:
        await ctx.send('Not a proper gauntlet.')
