from openpyxl import load_workbook


def findingstats(num):
    """Find the players placement"""
    storednums = 'First,Second,Third,Fourth,Fifth,Sixth,Seventh,Eighth,Ninth,Tenth,' \
                 'Eleventh,Twelfth,Thirteenth,Fourteenth,Fifteenth,Sixteenth,Seventeenth,' \
                 'Eighteenth,Nineteenth,Twentieth,Twenty-first,Twenty-second,Twenty-third,Twenty-fourth,' \
                 'Twenty-fifth,Twenty-sixth,Twenty-seventh,Twenty-eighth,Twenty-ninth,Thirtieth,Thirty-first,' \
                 'Thirty-second,Thirty-third,Thirty-fourth,Thirty-fifth,Thirty-sixth,Thirty-seventh,Thirty-eighth,' \
                 'Thirty-ninth,Fortieth,Forty-first,Forty-second,Forty-third,Forty-fourth,Forty-fifth,Forty-sixth,' \
                 'Forty-seventh,Forty-eighth,Forty-ninth,Fiftieth'
    splitnums = storednums.strip().upper().split(',')
    counting = 0
    for n in num:
        for s in splitnums:
            if n == s:
                return num[counting:counting + 4]
        counting += 1
    return 0


def checkchannels(channels, check):
    """Check for a specific channel"""
    for c in channels:
        if str(c) == check:
            return True
    return False


def check_s1(check):
    """Opens temp scrim role exel sheet, reads and writes information"""
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    count = 1
    for row in ws.iter_rows(values_only=True):
        if str(row[0]) == str(check):
            lb.close()
            return True
        else:
            pass
        count += 1
    ws[f'A{count}'] = str(check)
    lb.save('scrimrole.xlsx')
    lb.close()
    return False


def check_s2(check):
    """Opens temp scrim role exel sheet, reads and writes information"""
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    count = 1
    for row in ws.iter_rows(values_only=True):
        if row[1] == str(check):
            lb.close()
            return True
        else:
            pass
        count += 1
    ws[f'B{count}'] = str(check)
    lb.save('scrimrole.xlsx')
    lb.close()
    return False


def check_s3(check):
    """Opens temp scrim role exel sheet, reads and writes information"""
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    count = 1
    for row in ws.iter_rows(values_only=True):
        if row[2] == str(check):
            lb.close()
            return True
        else:
            pass
        count += 1
    ws[f'C{count}'] = str(check)
    lb.save('scrimrole.xlsx')
    lb.close()
    return False