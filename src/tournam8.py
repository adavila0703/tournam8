from discord import ChannelType
from discord.ext import commands
from discord.ext.commands import Bot, has_permissions, CommandNotFound
import discord.utils
import os
import cv2
import pytesseract
import requests
import json
from requests.auth import HTTPBasicAuth
import time
import random
from openpyxl import load_workbook

'''Global variables'''
bot = Bot(command_prefix='!')
bot.remove_command('help')

tournyname = ''
tournystart = False
marladuelwins = 0

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'


def get_token():
    with open('token.txt', 'r') as g:
        lines = g.readlines()
        return lines[0].strip()


''' Non bot commands '''


def findingstats(num):
    storednums = 'First,Second,Third,Fourth,Fifth,Sixth,Seventh,Eighth,Ninth,Tenth,' \
                 'Eleventh,Twelfth,Thirteenth,Fourteenth,Fifteenth,Sixteenth,Seventeenth,' \
                 'Eighteenth,Nineteenth,Twentieth,Twenty-first,Twenty-second,Twenty-third,Twenty-fourth,' \
                 'Twenty-fifth,Twenty-sixth,Twenty-seventh,Twenty-eighth,Twenty-ninth,Thirtieth,Thirty-first,' \
                 'Thirty-second,Thirty-third,Thirty-fourth,Thirty-fifth,Thirty-sixth,Thirty-seventh,Thirty-eighth,' \
                 'Thirty-ninth,Fortieth,Forty-first,Forty-second,Forty-third,Forty-fourth,Forty-fifth,Forty-sixth,' \
                 'Forty-seventh,Forty-eighth,Forty-ninth,Fiftieth'
    splitnums = storednums.strip().upper().split(',')
    counting = 0
    for n in num:
        for s in splitnums:
            if n == s:
                return num[counting:counting + 4]
        counting += 1
    return 0


def checkchannels(channels, check):
    for c in channels:
        if str(c) == check:
            return True
    return False


'''Bot functions that can be accepted by messages'''


@bot.event
async def on_ready():
    print('We have logged in as {0.user}'.format(bot))


@bot.event
async def on_message(message):
    global gamecount
    elduel = ['Fire', 'Toxic', 'Wind', 'Ice', 'Lightning']

    # regular discord message fun
    if message.author == bot.user:
        return

    if 'test' in message.content.lower():
        pass

    if 'marla' in message.content.lower():
        await message.channel.send('Yes?')

    if 'best rune' in message.content.lower():
        await message.channel.send('Blink...')

    if 'blink' in message.content.lower() or 'blinked' in message.content.lower():
        await message.channel.send('RIP Blink :(')

    if 'pink marla' in message.content.lower() or 'pinkmarla' in message.content.lower():
        await message.channel.send('Pink Marla is an impostor!')

    if 'marla who created you?' in message.content.lower() or 'marla who created you' in message.content.lower() \
            or 'who created marla' in message.content.lower() or 'who made marla' in message.content.lower() \
            or 'who made marla bot' in message.content.lower() or 'who created marla bot' in message.content.lower() \
            or 'marla who made you?' in message.content.lower() or 'marla who made you' in message.content.lower():
        await message.channel.send('I was created by marley-EE!')

    if 'where is marley' in message.content.lower() or 'where is marley-ee' in message.content.lower() \
            or 'wheres marley' in message.content.lower() or "where's marley" in message.content.lower() \
            or 'wheres marley-ee' in message.content.lower() or "where's marley-ee" in message.content.lower():
        await message.channel.send('marley-EE is at the grocery store...')

    if 'tell me a joke' in message.content.lower():
        print('nothing')

    # OCR this is where the bot reads the values from a screenshots and sends the data to a database
    if message.attachments:
        if tournystart == True:
            if str(message.channel) == str(message.author).split('#')[0].lower():
                await message.attachments[0].save(str(message.author).split('#')[0] + '.png')
                image = cv2.imread(str(message.author).split('#')[0] + '.png', 0)
                thresh = cv2.threshold(image, 150, 255, cv2.THRESH_BINARY_INV)[1]
                stats = findingstats(pytesseract.image_to_string(thresh, lang='eng', config='--psm 12', nice=1).split())
                try:
                    await message.channel.send(f'Player: {message.author} \nPlace: {stats[0]} \nExiles: {stats[1]}'
                                               f'\nAssists: {stats[2]} \nDamage: {stats[3]}')
                except:
                    pass
                os.remove(str(message.author).split('#')[0] + '.png')
                # if str(message.attachments).split()[3].split("'")[1].endswith('.png'):
            else:
                pass
        else:
            pass
    await bot.process_commands(message)


# function that changes user role on voice channel join and leave
@bot.event
async def on_voice_state_update(member, before, after):
    try:
        if str(after.channel) == 'Scrimming-VC':
            role = discord.utils.get(member.guild.roles, name="Player")
            await member.add_roles(role)
        elif str(before.channel) == 'Scrimming-VC':
            role = discord.utils.get(member.guild.roles, name="Player")
            await member.remove_roles(role)
    except AttributeError:
        pass


''' Bot commands '''


# element duel game!
@bot.command()
async def duel(ctx, choice):
    elduel = ['fire', 'toxic', 'wind', 'ice', 'lightning']
    lb = load_workbook(filename='duelstats.xlsx')
    ws = lb.active
    sheet_ranges = lb['Sheet']
    count = 1
    marlawins = 0

    if choice.lower() == 'fire' or choice.lower() == 'toxic' or choice.lower() == 'wind' or choice.lower() == 'ice' or \
            choice.lower() == 'lightning':
        marlapick = random.choice(elduel)
        await ctx.send(f"I choose {marlapick}!")

        if marlapick == 'fire':
            if choice.lower() == 'ice' or choice.lower() == 'toxic':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'fire':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'toxic':
            if choice.lower() == 'wind' or choice.lower() == 'lightning':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'toxic':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'wind':
            if choice.lower() == 'lightning' or choice.lower() == 'fire':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'wind':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'ice':
            if choice.lower() == 'toxic' or choice.lower() == 'wind':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'ice':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'lightning':
            if choice.lower() == 'fire' or choice.lower() == 'ice':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'lightning':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

    elif choice.lower() == 'stone':
        await ctx.send('Stone beats everything and it is too easy to use, choose a different gauntlet!')
    else:
        await ctx.send('Not a proper gauntlet.')


# testing command
@bot.command()
@has_permissions(administrator=True)
async def testingpost(ctx):
    ply = ctx.author.id
    print(ply)
    url = "http://localhost:5000/records"

    payload = "{\r\n    \"data\":\r\n    {\r\n    \"type\":\"user\", \"attributes\":\r\n        {\r\n        " \
              "\"username\":\"test\", \r\n        \"discordname\":\"test\"\r\n\r\n        }\r\n    }\r\n} "
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Basic dGl0bzp0aXRv'
    }

    filter = '''{"name":"discord_id", "op":"eq", "val":"''' + "{a}".format(a=ply) + '''"}'''

    r = requests.get(url + f"?filter=[{filter}]", auth=HTTPBasicAuth('tito', 'tito'))
    read = json.loads(r.text)
    data = read['data']
    kills = 0
    assists = 0
    damage = 0

    for d in data:
        kills += int(d['attributes']['kills'])

    for d in data:
        assists += int(d['attributes']['assists'])

    for d in data:
        damage += int(d['attributes']['damage'])

    await ctx.send('Kills: {a}\nAssists: {b}\nTotal Damage:{c}'.format(a=kills, b=assists, c=damage))


# command to get stats from database using an api
@bot.command()
@has_permissions(administrator=True)
async def getstats(ctx, player):
    ply = player.split('!')[1].replace('>', '')
    url = "http://localhost:5000/records"

    payload = "{\r\n    \"data\":\r\n    {\r\n    \"type\":\"user\", \"attributes\":\r\n        {\r\n        " \
              "\"username\":\"test\", \r\n        \"discordname\":\"test\"\r\n\r\n        }\r\n    }\r\n} "
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Basic dGl0bzp0aXRv'
    }

    filter = '''{"name":"discord_id", "op":"eq", "val":"''' + "{a}".format(a=ply) + '''"}'''

    r = requests.get(url + f"?filter=[{filter}]", auth=HTTPBasicAuth('tito', 'tito'))
    read = json.loads(r.text)
    data = read['data']
    kills = 0
    assists = 0
    damage = 0

    for d in data:
        kills += int(d['attributes']['kills'])

    for d in data:
        assists += int(d['attributes']['assists'])

    for d in data:
        damage += int(d['attributes']['damage'])

    await ctx.send('Kills: {a}\nAssists: {b}\nTotal Damage:{c}'.format(a=kills, b=assists, c=damage))


# test
@bot.command()
@has_permissions(administrator=True)
async def testget(ctx):
    url = "http://localhost:5000/users"

    payload = "{\r\n    \"data\":\r\n    {\r\n    \"type\":\"tourns\", \r\n    \"attributes\":\r\n        {\r\n       " \
              " \"name\": \"element 4\"\r\n\r\n        }\r\n    }\r\n} "
    headers = {
        'Authorization': 'Basic dGl0bzp0aXRv',
        'Content-Type': 'application/json'
    }

    # response = requests.request("GET", url, headers=headers, data=payload)

    filter = '''{"name":"username", "op":"eq", "val":"marley"}'''
    r = requests.get(url + f"?filter=[{filter}]", auth=HTTPBasicAuth('tito', 'tito'))

    read = json.loads(r.text)
    # data = read['data']

    # for d in data:
    #     print(d['attributes']['kills'])
    # return None

    print(read)


# testing a post command
@bot.command()
@has_permissions(administrator=True)
async def testpost(ctx):
    url = "http://localhost:5000/users"

    payload = "{\r\n    \"data\":\r\n    {\r\n    \"type\":\"user\", \"attributes\":\r\n        {\r\n        " \
              "\"username\":\"rip\", \r\n        \"discordname\":\"hello\"\r\n\r\n        }\r\n    }\r\n} "
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Basic dGl0bzp0aXRv'
    }

    requests.request("POST", url, headers=headers, data=payload)
    return None


# testing command
@bot.command()
@has_permissions(administrator=True)
async def starttournament(ctx, name):
    global tournyname, tournystart
    file = open('players.txt', 'r')
    if tournystart == False:
        tournystart = True
        tournyname = name
        print(tournyname)
        await ctx.send(f'Starting tournment {tournyname}')
        for f in file.readlines():
            text_out = discord.utils.get(ctx.guild.text_channels, name=f.split('\n')[0])
            await text_out.send('Tournment is starting! Only the screenshot that relates to the game you are'
                                'currently playing.\n\n'
                                '---------- Game 1 ----------')

    else:
        await ctx.send('There is already a tournment in progress')


# command that makes a text channel for all users in a given voice chat
@bot.command()
@has_permissions(administrator=True)
async def makescreenshotchannel(ctx, vc_channel, category):
    await ctx.send('Executing !makescreenshotchannel')
    start_time = time.time()
    file = open('players.txt', 'w')
    channel = discord.utils.get(ctx.guild.voice_channels, name=vc_channel)
    get_text_channels = ctx.guild.text_channels
    if discord.utils.get(ctx.guild.categories, name=category) is None:
        await ctx.send('Category doesnt exist')
    else:
        for c in channel.members:
            for g in get_text_channels:
                if str(c).split('#')[0].lower() == str(g):
                    return None
                else:
                    pass
            await ctx.guild.create_text_channel(str(c).split('#')[0],
                                                category=discord.utils.get(ctx.guild.categories, name=category))
            name = str(c).split('#')[0].lower()
            file.writelines(name + '\n')
            nameout = c.mention
            text_out = discord.utils.get(ctx.guild.text_channels, name=name)
            try:
                await text_out.send(
                    f'Welcome {nameout}! This is the channel where you will be posting your screenshots,'
                    f'dont forget!\n\nOne of my major features is the ability to read the'
                    f' information from your screenshot, make sure you get a good picture of your stats'
                    f' after your game, the better the picture, the more accurate I will be!'
                    f'\n\nBelow is an and example of the screen you need to capture after your '
                    f'Spellbreak '
                    f'match!\n\nGood luck today Breaker!\n\n -Marla')
                await text_out.send(file=discord.File('marley.png'))
                await text_out.send('Tip: ALT + PRINTSCREEN will take a picture of the monitor your mouse is currently'
                                    'active on. \n\nCTL + v into the discord message will send the picture you just '
                                    'took.')
            except AttributeError:
                await ctx.send('Cant accept special characters a players name.')
                pass
    await ctx.send(f'!makescreenshotchannel Completed! - Execution Time: {time.time() - start_time}s')


# OLD command to delete channels in a given category
@bot.command()
@commands.has_permissions(administrator=True)
async def old_deletechannels(ctx, category):
    file = open('players.txt', 'r')
    await ctx.send('Executing !deletechannels')
    start_time = time.time()
    category = discord.utils.get(ctx.guild.categories, name=category)
    for c in category.text_channels:
        await c.delete()
    await ctx.send(f'!deletechannels Completed! - Execution Time: {time.time() - start_time}s')


# NEW command to delete channels in a given category
@bot.command()
@commands.has_permissions(administrator=True)
async def deletechannels(ctx):
    file = open('players.txt', 'r')
    await ctx.send('Executing !deletechannels')
    start_time = time.time()
    for a in file.readlines():
        channame = a.strip().replace('\n', '')
        channel = discord.utils.get(ctx.guild.text_channels, name=channame)
        await channel.delete()
    await ctx.send(f'!deletechannels Completed! - Execution Time: {time.time() - start_time}s')


# command to give roles to players in a given voice chat
@bot.command()
@has_permissions(manage_roles=True)
async def giverole(ctx, role_name, vc_channel):
    await ctx.send('Executing !giverole')
    start_time = time.time()
    fetch = discord.utils.get(ctx.guild.voice_channels, name=vc_channel)
    user = fetch.members
    role = discord.utils.get(ctx.guild.roles, name=role_name)
    for u in user:
        await u.add_roles(role)
    await ctx.send(f'!giverole Completed! - Execution Time: {time.time() - start_time}s')

@bot.command()
@has_permissions(manage_roles=True)
async def test(ctx):
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    print(ws['A1'].value)

# global functions mean for scrim role function

def check_s1(check):
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    count = 1
    for row in ws.iter_rows(values_only=True):
        if str(row[0]) == str(check):
            lb.close()
            return True
        else:
            pass
        count += 1
    ws[f'A{count}'] = str(check)
    lb.save('scrimrole.xlsx')
    lb.close()
    return False

def check_s2(check):
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    count = 1
    for row in ws.iter_rows(values_only=True):
        if row[1] == str(check):
            lb.close()
            return True
        else:
            pass
        count += 1
    ws[f'B{count}'] = str(check)
    lb.save('scrimrole.xlsx')
    lb.close()
    return False

def check_s3(check):
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    count = 1
    for row in ws.iter_rows(values_only=True):
        if row[2] == str(check):
            lb.close()
            return True
        else:
            pass
        count += 1
    ws[f'C{count}'] = str(check)
    lb.save('scrimrole.xlsx')
    lb.close()
    return False

# rando
@bot.command()
@has_permissions(administrator=True)
async def updatescrimroles(ctx):
    await ctx.send('Executing !updatescrimroles')
    start_time = time.time()
    channels = ctx.guild.voice_channels
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')

    for c in channels:
        for a in c.members:
            if check_s1(a) == False:
                await a.add_roles(s1)
            elif check_s2(a) == False:
                await a.add_roles(s2)
                await a.remove_roles(s1)
            elif check_s3(a) == False:
                await a.add_roles(s3)
                await a.remove_roles(s2)
            else:
                pass

    await ctx.send(f'!updatescrimroles Completed! - Execution Time: {time.time() - start_time}s')

# command to update a role in a given voice chat
@bot.command()
@has_permissions(manage_roles=True)
async def old_updatescrimroles(ctx, vc_channel):
    await ctx.send('Executing !updatescrimroles')
    start_time = time.time()
    fetch = discord.utils.get(ctx.guild.voice_channels, name=vc_channel)
    user = fetch.members
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')
    for u in user:
        for r in u.roles:
            if r == s2:
                await u.add_roles(s3)
            elif r == s1:
                await u.add_roles(s2)
            else:
                await u.add_roles(s1)

    await ctx.send(f'!updatescrimroles Completed! - Execution Time: {time.time() - start_time}s')

@bot.command()
@has_permissions(administrator=True)
async def removeallscrimroles(ctx):
    await ctx.send('Executing !removeallscrimroles')
    start_time = time.time()
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')
    count = 1
    for row in ws.iter_rows(values_only=True):
        try:
            user = discord.utils.get(ctx.guild.members, name=str(row[2].split('#')[0]))
            await user.remove_roles(s3)
        except AttributeError:
            pass
    for row in ws.iter_rows(values_only=True):
        try:
            user = discord.utils.get(ctx.guild.members, name=str(row[1].split('#')[0]))
            await user.remove_roles(s2)
        except AttributeError:
            pass
    for row in ws.iter_rows(values_only=True):
        try:
            user = discord.utils.get(ctx.guild.members, name=str(row[0].split('#')[0]))
            await user.remove_roles(s1)
        except AttributeError:
            pass

    for row in ws.iter_rows(values_only=True):
        ws[f'A{count}'] = None
        ws[f'B{count}'] = None
        ws[f'C{count}'] = None
        count += 1

    ws[f'A1'] = 'null'
    ws[f'B1'] = 'null'
    ws[f'C1'] = 'null'
    lb.save('scrimrole.xlsx')
    await ctx.send(f'!removeallscrimroles Completed! - Execution Time: {time.time() - start_time}s')


# command to mass remove a specific role to all users in the discord
@bot.command()
@has_permissions(administrator=True)
async def old_removeallscrimroles(ctx):
    await ctx.send('Executing !removeallscrimroles')
    start_time = time.time()
    fetch = ctx.guild.members
    s1 = discord.utils.get(ctx.guild.roles, name='Scrimmed 1x')
    s2 = discord.utils.get(ctx.guild.roles, name='Scrimmed 2x')
    s3 = discord.utils.get(ctx.guild.roles, name='Scrimmed 3x')
    for f in fetch:
        for r in f.roles:
            if r == s1 or r == s2 or r == s3:
                await f.remove_roles(s1)
                await f.remove_roles(s2)
                await f.remove_roles(s3)
    await ctx.send(f'!removeallscrimroles Completed! - Execution Time: {time.time() - start_time}s')


# help command that displays all the available commands
@bot.command()
@has_permissions(manage_roles=True)
async def help(member):
    await member.send("```                            Help Menu\n"
                      "\nHow to use: !func(p1, p2) will be written out like !func p1 p2, for spaces use"
                      "quotes!\n\n"
                      "-!makescreenshotchannel(vc_channel, category) - Creates a text channel for all the users "
                      "inside a specified voice channel\n\n"
                      "-!deletechannels(category) - Deletes all text channels in a selected category.\n\n"
                      "-!giverole(role_name, vc_channel) - Give everyone in a specified voice chat a role of "
                      "choice.\n\n "
                      "-!updatescrimroles - Gives a user the 'scrim role'(x1, x2, x3) Example: If they "
                      "have x1 and you run the function, then x1 will be deleted and x2 will be added.\n\n"
                      "-!removeallscrimroles - Removes all scrim roles from everyone who owns the role.\n\n"
                      "** Marla bot was created by marley-EE **```")


''' Bot command errors '''


@makescreenshotchannel.error
async def makescreenshotchannel_error(ctx, error):
    if isinstance(error, commands.MissingRequiredArgument):
        await ctx.send('Error: You are missing your arguments')
    if isinstance(error, commands.CommandInvokeError):
        await ctx.send('Error: -No users in the channel\n -Category or channel doesnt exist\n '
                       '-Bot doesnt have permissions')


@deletechannels.error
async def deletechannels_error(ctx, error):
    if isinstance(error, commands.MissingRequiredArgument):
        await ctx.send('Error: You are missing your arguments')
    if isinstance(error, commands.CommandInvokeError):
        await ctx.send('Error: No channels or category doesnt exist.')


# @giverole.error
# async def giverole_error(ctx, error):
#     if isinstance(error, commands.MissingRequiredArgument):
#         await ctx.send('Error: You are missing your arguments')
#     if isinstance(error, commands.CommandInvokeError):
#         await ctx.send('Error: No channels or category doesnt exist.')

@updatescrimroles.error
async def updatescrimroles_error(ctx, error):
    if isinstance(error, commands.MissingRequiredArgument):
        await ctx.send('Error: You are missing your arguments')
    if isinstance(error, commands.CommandInvokeError):
        await ctx.send('Error: Role might not exist.')


@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, CommandNotFound):
        await ctx.send('Error: Command not found or spelled incorrectly')
        return
    raise error


bot.run(get_token())
