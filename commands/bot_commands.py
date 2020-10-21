from bot.bot import bot
from openpyxl import load_workbook
import random
from discord.ext.commands import has_permissions
import requests
import json
from requests.auth import HTTPBasicAuth

# element duel game!
@bot.command()
async def duel(ctx, choice):
    elduel = ['fire', 'toxic', 'wind', 'ice', 'lightning']
    lb = load_workbook(filename='duelstats.xlsx')
    ws = lb.active
    sheet_ranges = lb['Sheet']
    count = 1
    marlawins = 0

    if choice.lower() == 'fire' or choice.lower() == 'toxic' or choice.lower() == 'wind' or choice.lower() == 'ice' or \
            choice.lower() == 'lightning':
        marlapick = random.choice(elduel)
        await ctx.send(f"I choose {marlapick}!")

        if marlapick == 'fire':
            if choice.lower() == 'ice' or choice.lower() == 'toxic':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'fire':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'toxic':
            if choice.lower() == 'wind' or choice.lower() == 'lightning':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'toxic':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'wind':
            if choice.lower() == 'lightning' or choice.lower() == 'fire':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'wind':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'ice':
            if choice.lower() == 'toxic' or choice.lower() == 'wind':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'ice':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

        elif marlapick == 'lightning':
            if choice.lower() == 'fire' or choice.lower() == 'ice':
                for row in ws.iter_rows(values_only=True):
                    if row[0] == 'Marla':
                        sheet_ranges[f'B{count}'].value += 1
                        marlawins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                    else:
                        pass
                await ctx.send(f'I win! {marlapick.capitalize()} beats {choice.capitalize()}. Marla wins: {marlawins}')
                return None
            elif choice.lower() == 'lightning':
                await ctx.send(f'Draw!')
                return None
            else:
                for row in ws.iter_rows(values_only=True):
                    if row[0] == str(ctx.author):
                        sheet_ranges[f'B{count}'].value += 1
                        yourwins = sheet_ranges[f'B{count}'].value
                        lb.save('duelstats.xlsx')
                        await ctx.send(
                            f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')
                        return None
                    else:
                        pass
                    count += 1
                ws[f'A{count}'] = str(ctx.author)
                ws[f'B{count}'] = 0
                yourwins = sheet_ranges[f'B{count}'].value
                lb.save('duelstats.xlsx')
                await ctx.send(f'You win! {choice.capitalize()} beats {marlapick.capitalize()}. Your wins: {yourwins}')

    elif choice.lower() == 'stone':
        await ctx.send('Stone beats everything and it is too easy to use, choose a different gauntlet!')
    else:
        await ctx.send('Not a proper gauntlet.')


# testing command
@bot.command()
@has_permissions(administrator=True)
async def testingpost(ctx):
    ply = ctx.author.id
    print(ply)
    url = "http://localhost:5000/records"

    payload = "{\r\n    \"data\":\r\n    {\r\n    \"type\":\"user\", \"attributes\":\r\n        {\r\n        " \
              "\"username\":\"test\", \r\n        \"discordname\":\"test\"\r\n\r\n        }\r\n    }\r\n} "
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Basic dGl0bzp0aXRv'
    }

    filter = '''{"name":"discord_id", "op":"eq", "val":"''' + "{a}".format(a=ply) + '''"}'''

    r = requests.get(url + f"?filter=[{filter}]", auth=HTTPBasicAuth('tito', 'tito'))
    read = json.loads(r.text)
    data = read['data']
    kills = 0
    assists = 0
    damage = 0

    for d in data:
        kills += int(d['attributes']['kills'])

    for d in data:
        assists += int(d['attributes']['assists'])

    for d in data:
        damage += int(d['attributes']['damage'])

    await ctx.send('Kills: {a}\nAssists: {b}\nTotal Damage:{c}'.format(a=kills, b=assists, c=damage))


# command to get stats from database using an api
@bot.command()
@has_permissions(administrator=True)
async def getstats(ctx, player):
    ply = player.split('!')[1].replace('>', '')
    file = open('element-info.txt', 'r')
    out = file.readlines()
    url = f"http://{out[2].strip()}:{out[3].strip()}/records"

    filter = '''{"name":"discord_id", "op":"eq", "val":"''' + "{a}".format(a=ply) + '''"}'''

    r = requests.get(url + f"?filter=[{filter}]", auth=HTTPBasicAuth(f'{out[0].strip()}', f'{out[1].strip()}'))
    file.close()
    read = json.loads(r.text)
    data = read['data']
    kills = 0
    assists = 0
    damage = 0

    for d in data:
        kills += int(d['attributes']['kills'])

    for d in data:
        assists += int(d['attributes']['assists'])

    for d in data:
        damage += int(d['attributes']['damage'])

    await ctx.send('Kills: {a}\nAssists: {b}\nTotal Damage:{c}'.format(a=kills, b=assists, c=damage))


# test
@bot.command()
@has_permissions(administrator=True)
async def testget(ctx):
    url = "http://localhost:5000/users"

    payload = "{\r\n    \"data\":\r\n    {\r\n    \"type\":\"tourns\", \r\n    \"attributes\":\r\n        {\r\n       " \
              " \"name\": \"element 4\"\r\n\r\n        }\r\n    }\r\n} "
    headers = {
        'Authorization': 'Basic dGl0bzp0aXRv',
        'Content-Type': 'application/json'
    }

    # response = requests.request("GET", url, headers=headers, data=payload)

    filter = '''{"name":"username", "op":"eq", "val":"marley"}'''
    r = requests.get(url + f"?filter=[{filter}]", auth=HTTPBasicAuth('tito', 'tito'))

    read = json.loads(r.text)
    # data = read['data']
    #
    # score = 0
    # discord_name = 'marley'
    # for d in data:
    #     d['attributes']['score'] += score
    # return None
    #
    # print(discord_name + '-' + score)
    #
    # print(read)


# testing a post command
@bot.command()
@has_permissions(administrator=True)
async def testpost(ctx):
    url = "http://localhost:5000/users"

    payload = "{\r\n    \"data\":\r\n    {\r\n    \"type\":\"user\", \"attributes\":\r\n        {\r\n        " \
              "\"username\":\"rip\", \r\n        \"discordname\":\"hello\"\r\n\r\n        }\r\n    }\r\n} "
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Basic dGl0bzp0aXRv'
    }

    requests.request("POST", url, headers=headers, data=payload)
    return None


# testing command
@bot.command()
@has_permissions(administrator=True)
async def starttournament(ctx, name):
    global tournyname, tournystart
    file = open('players.txt', 'r')
    if tournystart == False:
        tournystart = True
        tournyname = name
        print(tournyname)
        await ctx.send(f'Starting tournment {tournyname}')
        for f in file.readlines():
            text_out = discord.utils.get(ctx.guild.text_channels, name=f.split('\n')[0])
            await text_out.send('Tournment is starting! Only the screenshot that relates to the game you are'
                                'currently playing.\n\n'
                                '---------- Game 1 ----------')

    else:
        await ctx.send('There is already a tournment in progress')


# command that makes a text channel for all users in a given voice chat
@bot.command()
@has_permissions(administrator=True)
async def makescreenshotchannel(ctx, vc_channel, category):
    await ctx.send('Executing !makescreenshotchannel')
    start_time = time.time()
    file = open('players.txt', 'w')
    channel = discord.utils.get(ctx.guild.voice_channels, name=vc_channel)
    get_text_channels = ctx.guild.text_channels
    if discord.utils.get(ctx.guild.categories, name=category) is None:
        await ctx.send('Category doesnt exist')
    else:
        for c in channel.members:
            for g in get_text_channels:
                if str(c).split('#')[0].lower() == str(g):
                    return None
                else:
                    pass
            await ctx.guild.create_text_channel(str(c).split('#')[0],
                                                category=discord.utils.get(ctx.guild.categories, name=category))
            name = str(c).split('#')[0].lower()
            file.writelines(name + '\n')
            nameout = c.mention
            text_out = discord.utils.get(ctx.guild.text_channels, name=name)
            try:
                await text_out.send(
                    f'Welcome {nameout}! This is the channel where you will be posting your screenshots,'
                    f'dont forget!\n\nOne of my major features is the ability to read the'
                    f' information from your screenshot, make sure you get a good picture of your stats'
                    f' after your game, the better the picture, the more accurate I will be!'
                    f'\n\nBelow is an and example of the screen you need to capture after your '
                    f'Spellbreak '
                    f'match!\n\nGood luck today Breaker!\n\n -Marla')
                await text_out.send(file=discord.File('marley.png'))
                await text_out.send('Tip: ALT + PRINTSCREEN will take a picture of the monitor your mouse is currently'
                                    'active on. \n\nCTL + v into the discord message will send the picture you just '
                                    'took.')
            except AttributeError:
                await ctx.send('Cant accept special characters a players name.')
                pass
    await ctx.send(f'!makescreenshotchannel Completed! - Execution Time: {time.time() - start_time}s')


# OLD command to delete channels in a given category
@bot.command()
@commands.has_permissions(administrator=True)
async def old_deletechannels(ctx, category):
    file = open('players.txt', 'r')
    await ctx.send('Executing !deletechannels')
    start_time = time.time()
    category = discord.utils.get(ctx.guild.categories, name=category)
    for c in category.text_channels:
        await c.delete()
    await ctx.send(f'!deletechannels Completed! - Execution Time: {time.time() - start_time}s')


# NEW command to delete channels in a given category
@bot.command()
@commands.has_permissions(administrator=True)
async def deletechannels(ctx):
    file = open('players.txt', 'r')
    await ctx.send('Executing !deletechannels')
    start_time = time.time()
    for a in file.readlines():
        channame = a.strip().replace('\n', '')
        channel = discord.utils.get(ctx.guild.text_channels, name=channame)
        await channel.delete()
    await ctx.send(f'!deletechannels Completed! - Execution Time: {time.time() - start_time}s')


# command to give roles to players in a given voice chat
@bot.command()
@has_permissions(manage_roles=True)
async def giverole(ctx, role_name, vc_channel):
    await ctx.send('Executing !giverole')
    start_time = time.time()
    fetch = discord.utils.get(ctx.guild.voice_channels, name=vc_channel)
    user = fetch.members
    role = discord.utils.get(ctx.guild.roles, name=role_name)
    for u in user:
        await u.add_roles(role)
    await ctx.send(f'!giverole Completed! - Execution Time: {time.time() - start_time}s')


@bot.command()
@has_permissions(manage_roles=True)
async def test(ctx):
    lb = load_workbook(filename='scrimrole.xlsx')
    ws = lb.active
    print(ws['A1'].value)